from contextlib import asynccontextmanager
import csv
import io
import json
import re
from datetime import date
from decimal import Decimal
from typing import Annotated

from fastapi import Depends, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from pydantic import BaseModel, EmailStr

from app.auth import (
    create_access_token,
    ensure_offering_access,
    generate_temporary_password,
    get_current_user,
    hash_password,
    is_valid_password,
    require_offering_access,
    require_permission,
    verify_password,
)
from app.config import get_settings
from app.db import fetch_all, fetch_one, get_conn
from app.migrations import run_migrations
from app.seed import seed_demo_data
from app.services.calculation import split_weight
from app.services.grade_import import parse_mark, weighted_score
from app.services.handbook import HandbookImportError, fetch_handbook


@asynccontextmanager
async def lifespan(app: FastAPI):
    run_migrations()
    seed_demo_data()
    yield


app = FastAPI(title="MCS07 Academic Performance API", version="0.1.0", lifespan=lifespan)
settings = get_settings()

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class PasswordChangeRequest(BaseModel):
    current_password: str
    new_password: str


class AdminUserCreate(BaseModel):
    staff_id: str
    full_name: str
    email: EmailStr
    role_name: str


class AdminUserUpdate(BaseModel):
    is_active: bool | None = None
    role_name: str | None = None


class AdminUserBulkCreate(BaseModel):
    users: list[AdminUserCreate]


class MappingUpdate(BaseModel):
    offering_id: int
    mappings: list[dict[str, int]]


class HandbookImportConfirmation(BaseModel):
    handbook_import_id: int


class AdminSemesterCreate(BaseModel):
    year: int
    period: str
    start_date: date | None = None
    end_date: date | None = None
    status: str = "planning"


class AdminSemesterUpdate(BaseModel):
    start_date: date | None = None
    end_date: date | None = None
    status: str


class AdminOfferingCreate(BaseModel):
    semester_id: int
    program_ids: list[int]
    unit_code: str
    unit_name: str
    coordinator_id: int
    lecturer_ids: list[int] = []
    status: str = "draft"


class AdminOfferingUpdate(BaseModel):
    unit_code: str
    unit_name: str
    program_ids: list[int]
    coordinator_id: int
    lecturer_ids: list[int] = []
    status: str
    replacement_unit_code: str | None = None
    replacement_unit_name: str | None = None


@app.get("/api/health")
def health():
    return {"status": "ok", "service": "backend"}


@app.get("/api/db-health")
def db_health():
    row = fetch_one("SELECT 1 AS ok")
    return {"status": "ok", "database": row["ok"] == 1}


@app.post("/api/auth/login")
def login(payload: LoginRequest):
    user = fetch_one(
        """
        SELECT u.user_id, u.staff_id, u.full_name, u.email, u.password_hash, u.is_active,
               u.must_change_password, r.role_name, r.permission_level
        FROM app_user u
        JOIN role r ON r.role_id = u.role_id
        WHERE u.email = %s
        """,
        (str(payload.email).lower(),),
    )
    if not user or not user["is_active"] or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    public_user = {
        key: user[key]
        for key in ["user_id", "staff_id", "full_name", "email", "must_change_password", "role_name", "permission_level"]
    }
    return {
        "access_token": create_access_token(user),
        "token_type": "bearer",
        "user": public_user,
    }


@app.post("/api/auth/change-password")
def change_password(
    payload: PasswordChangeRequest,
    user: Annotated[dict, Depends(get_current_user)],
):
    user_with_password = fetch_one(
        "SELECT password_hash FROM app_user WHERE user_id = %s",
        (user["user_id"],),
    )
    if not user_with_password or not verify_password(payload.current_password, user_with_password["password_hash"]):
        raise HTTPException(status_code=401, detail="Current password is incorrect")
    if not is_valid_password(payload.new_password):
        raise HTTPException(status_code=422, detail="New password must be at least 12 characters")

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE app_user SET password_hash = %s, must_change_password = FALSE WHERE user_id = %s",
                (hash_password(payload.new_password), user["user_id"]),
            )
    return {"status": "changed"}


@app.get("/api/me")
def me(user: Annotated[dict, Depends(get_current_user)]):
    return {"user": user}


@app.get("/api/offerings")
def offerings(user: Annotated[dict, Depends(get_current_user)]):
    query = """
        SELECT
            o.offering_id,
            u.unit_code,
            u.unit_name,
            ARRAY_AGG(DISTINCT p.program_code ORDER BY p.program_code) AS program_codes,
            ARRAY_AGG(DISTINCT p.program_name ORDER BY p.program_name) AS program_names,
            s.year,
            s.period,
            o.coordinator_id,
            o.handbook_url,
            o.last_scraped_at
        FROM unit_offering o
        JOIN unit u ON u.unit_id = o.unit_id
        JOIN offering_program op ON op.offering_id = o.offering_id
        JOIN program p ON p.program_id = op.program_id
        JOIN semester s ON s.semester_id = o.semester_id
    """
    if user["role_name"] == "coordinator":
        query += """
            WHERE o.coordinator_id = %s
               OR EXISTS (
                    SELECT 1 FROM offering_lecturer ol
                    WHERE ol.offering_id = o.offering_id AND ol.lecturer_id = %s
               )
        """
        params = (user["user_id"], user["user_id"])
    elif user["role_name"] == "lecturer":
        query += " WHERE EXISTS (SELECT 1 FROM offering_lecturer ol WHERE ol.offering_id = o.offering_id AND ol.lecturer_id = %s)"
        params = (user["user_id"],)
    elif user["role_name"] == "management":
        params = None
    else:
        raise HTTPException(status_code=403, detail="Unknown role")
    query += """
        GROUP BY o.offering_id, u.unit_code, u.unit_name, s.year, s.period,
                 o.coordinator_id, o.handbook_url, o.last_scraped_at
        ORDER BY s.year DESC, s.period, u.unit_code
    """
    rows = fetch_all(query, params)
    for row in rows:
        row["can_edit"] = user["role_name"] == "management" or row["coordinator_id"] == user["user_id"]
        del row["coordinator_id"]
    return {"offerings": rows}


def _handbook_offering(offering_id: int) -> dict:
    offering = fetch_one(
        """
        SELECT o.offering_id, o.unit_id, u.unit_code, s.year, s.period, o.handbook_location
        FROM unit_offering o
        JOIN unit u ON u.unit_id = o.unit_id
        JOIN semester s ON s.semester_id = o.semester_id
        WHERE o.offering_id = %s
        """,
        (offering_id,),
    )
    if not offering:
        raise HTTPException(status_code=404, detail="Offering not found")
    return offering


@app.post("/api/offerings/{offering_id}/handbook-import")
def create_handbook_import(
    offering_id: int,
    user: Annotated[dict, Depends(require_permission(20))],
):
    ensure_offering_access(user, offering_id, min_permission_level=20)
    offering = _handbook_offering(offering_id)
    try:
        imported = fetch_handbook(
            offering["unit_code"],
            offering["year"],
            period=offering["period"],
            location=offering["handbook_location"],
        )
    except HandbookImportError as exc:
        raise HTTPException(status_code=502, detail="Could not import the public Monash Handbook record") from exc

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO handbook_import_snapshot (offering_id, source_url, handbook_version, payload)
                VALUES (%s, %s, %s, %s::jsonb)
                RETURNING handbook_import_id, source_url, handbook_version, status, imported_at, payload
                """,
                (
                    offering_id,
                    imported["source_url"],
                    imported["payload"].get("handbook_version"),
                    json.dumps(imported["payload"]),
                ),
            )
            snapshot = cur.fetchone()
    return {"import": snapshot}


@app.get("/api/offerings/{offering_id}/handbook-import")
def latest_handbook_import(
    offering_id: int,
    user: Annotated[dict, Depends(require_offering_access(20))],
):
    snapshot = fetch_one(
        """
        SELECT handbook_import_id, source_url, handbook_version, status, imported_at, confirmed_at, payload
        FROM handbook_import_snapshot
        WHERE offering_id = %s
        ORDER BY handbook_import_id DESC
        LIMIT 1
        """,
        (offering_id,),
    )
    return {"import": snapshot}


@app.post("/api/offerings/{offering_id}/handbook-import/confirm")
def confirm_handbook_import(
    offering_id: int,
    confirmation: HandbookImportConfirmation,
    user: Annotated[dict, Depends(require_permission(20))],
):
    ensure_offering_access(user, offering_id, min_permission_level=20)
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT handbook_import_id, source_url, payload
                FROM handbook_import_snapshot
                WHERE handbook_import_id = %s AND offering_id = %s AND status = 'draft'
                FOR UPDATE
                """,
                (confirmation.handbook_import_id, offering_id),
            )
            snapshot = cur.fetchone()
            if not snapshot:
                raise HTTPException(status_code=404, detail="Draft Handbook import not found")

            cur.execute("SELECT EXISTS(SELECT 1 FROM student_grade WHERE offering_id = %s) AS has_grades", (offering_id,))
            if cur.fetchone()["has_grades"]:
                raise HTTPException(
                    status_code=409,
                    detail="Assessment setup cannot be replaced after grades exist. Record an amendment and recalculate before changing it.",
                )

            payload = snapshot["payload"]
            if not isinstance(payload, dict):
                raise HTTPException(status_code=422, detail="Draft Handbook import is invalid")

            cur.execute(
                "UPDATE unit_offering SET handbook_url = %s, last_scraped_at = CURRENT_TIMESTAMP WHERE offering_id = %s",
                (snapshot["source_url"], offering_id),
            )
            cur.execute(
                """
                UPDATE unit SET unit_name = %s, default_handbook_url = %s
                WHERE unit_id = (SELECT unit_id FROM unit_offering WHERE offering_id = %s)
                """,
                (payload["title"], snapshot["source_url"], offering_id),
            )

            ulo_ids: dict[str, int] = {}
            for ulo in payload["learning_outcomes"]:
                cur.execute(
                    """
                    INSERT INTO offering_ulo (
                        offering_id, ulo_code, description, source, handbook_reference, confirmed_by, confirmed_at
                    )
                    VALUES (%s, %s, %s, 'handbook', %s, %s, CURRENT_TIMESTAMP)
                    ON CONFLICT (offering_id, ulo_code) DO UPDATE
                    SET description = EXCLUDED.description,
                        handbook_reference = EXCLUDED.handbook_reference,
                        confirmed_by = EXCLUDED.confirmed_by,
                        confirmed_at = EXCLUDED.confirmed_at
                    WHERE offering_ulo.source = 'handbook'
                    RETURNING offering_ulo_id
                    """,
                    (offering_id, ulo["code"], ulo["description"], ulo.get("reference"), user["user_id"]),
                )
                row = cur.fetchone()
                if row is None:
                    cur.execute(
                        "SELECT offering_ulo_id FROM offering_ulo WHERE offering_id = %s AND ulo_code = %s",
                        (offering_id, ulo["code"]),
                    )
                    row = cur.fetchone()
                ulo_ids[ulo["code"]] = row["offering_ulo_id"]

            for assessment in payload["assessments"]:
                cur.execute(
                    "SELECT source FROM assessment WHERE offering_id = %s AND assessment_name = %s",
                    (offering_id, assessment["name"]),
                )
                existing = cur.fetchone()
                if existing and existing["source"] != "handbook":
                    raise HTTPException(
                        status_code=409,
                        detail=f"Manual assessment '{assessment['name']}' has the same name as the Handbook import.",
                    )

            cur.execute("DELETE FROM assessment WHERE offering_id = %s AND source = 'handbook'", (offering_id,))
            for order, assessment in enumerate(payload["assessments"], start=1):
                cur.execute(
                    """
                    INSERT INTO assessment (
                        offering_id, assessment_name, weight, max_mark, assessment_order, is_hurdle,
                        source, confirmed_by, confirmed_at
                    )
                    VALUES (%s, %s, %s, 100, %s, %s, 'handbook', %s, CURRENT_TIMESTAMP)
                    RETURNING assessment_id
                    """,
                    (
                        offering_id,
                        assessment["name"],
                        Decimal(assessment["weight"]),
                        order,
                        assessment["is_hurdle"],
                        user["user_id"],
                    ),
                )
                assessment_id = cur.fetchone()["assessment_id"]
                linked_ulo_ids = [ulo_ids[code] for code in assessment["ulo_codes"] if code in ulo_ids]
                for offering_ulo_id, allocated_weight in split_weight(Decimal(assessment["weight"]), linked_ulo_ids).items():
                    cur.execute(
                        """
                        INSERT INTO assessment_ulo (
                            offering_id, assessment_id, offering_ulo_id, source, is_confirmed,
                            allocated_weight, confirmed_by, confirmed_at
                        )
                        VALUES (%s, %s, %s, 'handbook', TRUE, %s, %s, CURRENT_TIMESTAMP)
                        """,
                        (offering_id, assessment_id, offering_ulo_id, allocated_weight, user["user_id"]),
                    )

            cur.execute(
                """
                UPDATE handbook_import_snapshot
                SET status = 'confirmed', confirmed_by = %s, confirmed_at = CURRENT_TIMESTAMP
                WHERE handbook_import_id = %s
                """,
                (user["user_id"], confirmation.handbook_import_id),
            )
    return {"status": "confirmed"}


@app.get("/api/dashboard")
def dashboard(user: Annotated[dict, Depends(require_offering_access())], offering_id: int = 1):
    offering = fetch_one(
        """
        SELECT o.offering_id, u.unit_code, u.unit_name, s.year, s.period,
               ARRAY_AGG(DISTINCT p.program_name ORDER BY p.program_name) AS program_names
        FROM unit_offering o
        JOIN unit u ON u.unit_id = o.unit_id
        JOIN offering_program op ON op.offering_id = o.offering_id
        JOIN program p ON p.program_id = op.program_id
        JOIN semester s ON s.semester_id = o.semester_id
        WHERE o.offering_id = %s
        GROUP BY o.offering_id, u.unit_code, u.unit_name, s.year, s.period
        """,
        (offering_id,),
    )
    if not offering:
        raise HTTPException(status_code=404, detail="Offering not found")

    los = fetch_all(
        """
        SELECT
            ou.offering_ulo_id,
            ou.ulo_code,
            ou.description,
            COALESCE(c.average_attainment_pct, 0) AS average_attainment_pct,
            COALESCE(c.pass_rate_pct, 0) AS pass_rate_pct,
            COALESCE(c.enrolled_count, 0) AS enrolled_count,
            COALESCE(c.achieved_count, 0) AS achieved_count
        FROM offering_ulo ou
        LEFT JOIN cohort_ulo_attainment c ON c.offering_ulo_id = ou.offering_ulo_id
        WHERE ou.offering_id = %s
        ORDER BY ou.ulo_code
        """,
        (offering_id,),
    )
    assessments = fetch_all(
        """
        SELECT
            a.assessment_id,
            a.assessment_name,
            a.weight,
            a.max_mark,
            a.is_hurdle,
            ARRAY_REMOVE(ARRAY_AGG(ou.ulo_code ORDER BY ou.ulo_code), NULL) AS covers
        FROM assessment a
        LEFT JOIN assessment_ulo au ON au.assessment_id = a.assessment_id
        LEFT JOIN offering_ulo ou ON ou.offering_ulo_id = au.offering_ulo_id
        WHERE a.offering_id = %s
        GROUP BY a.assessment_id
        ORDER BY a.assessment_order
        """,
        (offering_id,),
    )
    report = fetch_one(
        """
        SELECT report_id, ai_summary, coordinator_comment, is_finalized, generated_at
        FROM ai_report
        WHERE offering_id = %s
        ORDER BY generated_at DESC
        LIMIT 1
        """,
        (offering_id,),
    )
    stats = {
        "student_count": fetch_one("SELECT COUNT(*) AS count FROM enrollment WHERE offering_id = %s", (offering_id,))["count"],
        "lo_count": len(los),
        "at_risk_count": sum(1 for item in los if item["pass_rate_pct"] < 70),
        "student_at_risk_count": fetch_one(
            """
            SELECT COUNT(*) AS count FROM (
                SELECT enrollment_id
                FROM student_ulo_attainment
                WHERE offering_id = %s
                GROUP BY enrollment_id
                HAVING AVG(attainment_pct) < 50
            ) at_risk_students
            """,
            (offering_id,),
        )["count"],
    }
    return {"offering": offering, "stats": stats, "learning_outcomes": los, "assessments": assessments, "report": report}


@app.get("/api/mappings")
def mappings(user: Annotated[dict, Depends(require_offering_access())], offering_id: int = 1):
    ulos = fetch_all(
        "SELECT offering_ulo_id, ulo_code, description FROM offering_ulo WHERE offering_id = %s ORDER BY ulo_code",
        (offering_id,),
    )
    plos = fetch_all(
        """
        SELECT DISTINCT p.plo_id, p.plo_code, p.description
        FROM plo p
        JOIN offering_program op ON op.program_id = p.program_id
        WHERE op.offering_id = %s
        ORDER BY p.plo_code
        """,
        (offering_id,),
    )
    rows = fetch_all(
        """
        SELECT mapping_id, offering_ulo_id, plo_id, confirmed_at
        FROM ulo_plo_mapping
        WHERE offering_id = %s AND is_active = TRUE
        ORDER BY offering_ulo_id, plo_id
        """,
        (offering_id,),
    )
    return {"ulos": ulos, "plos": plos, "mappings": rows}


@app.put("/api/mappings")
def save_mappings(
    payload: MappingUpdate,
    user: Annotated[dict, Depends(require_permission(20))],
):
    ensure_offering_access(user, payload.offering_id, min_permission_level=20)
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE ulo_plo_mapping SET is_active = FALSE, removed_by = %s, removed_at = CURRENT_TIMESTAMP WHERE offering_id = %s",
                (user["user_id"], payload.offering_id),
            )
            for item in payload.mappings:
                cur.execute(
                    """
                    INSERT INTO ulo_plo_mapping (offering_id, offering_ulo_id, plo_id, confirmed_by)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT DO NOTHING
                    """,
                    (payload.offering_id, item["offering_ulo_id"], item["plo_id"], user["user_id"]),
                )
    return {"status": "saved"}


@app.get("/api/assessments")
def assessments(user: Annotated[dict, Depends(require_offering_access())], offering_id: int = 1):
    rows = fetch_all(
        """
        SELECT
            a.assessment_id,
            a.assessment_name,
            a.weight,
            a.max_mark,
            a.is_hurdle,
            a.source,
            ARRAY_REMOVE(ARRAY_AGG(ou.ulo_code ORDER BY ou.ulo_code), NULL) AS covers,
            ARRAY_REMOVE(ARRAY_AGG(au.allocated_weight ORDER BY ou.ulo_code), NULL) AS allocated_weights
        FROM assessment a
        LEFT JOIN assessment_ulo au ON au.assessment_id = a.assessment_id
        LEFT JOIN offering_ulo ou ON ou.offering_ulo_id = au.offering_ulo_id
        WHERE a.offering_id = %s
        GROUP BY a.assessment_id
        ORDER BY a.assessment_order
        """,
        (offering_id,),
    )
    return {"assessments": rows}


@app.get("/api/admin/users")
def admin_users(user: Annotated[dict, Depends(require_permission(30))]):
    rows = fetch_all(
        """
        SELECT u.user_id, u.staff_id, u.full_name, u.email, u.is_active, u.must_change_password,
               u.created_at, r.role_name, r.permission_level
        FROM app_user u
        JOIN role r ON r.role_id = u.role_id
        ORDER BY r.permission_level DESC, u.full_name
        """
    )
    return {"users": rows}


def _admin_user_values(payload: AdminUserCreate) -> tuple[str, str, str, str]:
    staff_id = payload.staff_id.strip()
    full_name = payload.full_name.strip()
    email = str(payload.email).lower()
    role_name = payload.role_name.strip().lower()
    if not re.fullmatch(r"\d{7}", staff_id):
        raise HTTPException(status_code=422, detail="Staff ID must be exactly seven digits")
    if len(full_name) < 3:
        raise HTTPException(status_code=422, detail="Full name is required")
    if not email.endswith("@monash.edu"):
        raise HTTPException(status_code=422, detail="Use a Monash staff email address")
    return staff_id, full_name, email, role_name


def _insert_admin_user(cur, values: tuple[str, str, str, str]) -> dict:
    staff_id, full_name, email, role_name = values
    cur.execute("SELECT role_id FROM role WHERE role_name = %s", (role_name,))
    role = cur.fetchone()
    if not role:
        raise HTTPException(status_code=422, detail="Role must be management, coordinator, or lecturer")
    cur.execute("SELECT 1 FROM app_user WHERE staff_id = %s OR email = %s", (staff_id, email))
    if cur.fetchone():
        raise HTTPException(status_code=409, detail="A staff account already uses that ID or email")

    temporary_password = generate_temporary_password()
    cur.execute(
        """
        INSERT INTO app_user (staff_id, full_name, email, password_hash, role_id, must_change_password)
        VALUES (%s, %s, %s, %s, %s, TRUE)
        RETURNING user_id
        """,
        (staff_id, full_name, email, hash_password(temporary_password), role["role_id"]),
    )
    user_id = cur.fetchone()["user_id"]
    cur.execute(
        """
        SELECT u.user_id, u.staff_id, u.full_name, u.email, u.is_active, u.must_change_password,
               u.created_at, r.role_name, r.permission_level
        FROM app_user u
        JOIN role r ON r.role_id = u.role_id
        WHERE u.user_id = %s
        """,
        (user_id,),
    )
    return {"user": cur.fetchone(), "temporary_password": temporary_password}


@app.post("/api/admin/users", status_code=201)
def create_admin_user(
    payload: AdminUserCreate,
    user: Annotated[dict, Depends(require_permission(30))],
):
    with get_conn() as conn:
        with conn.cursor() as cur:
            return _insert_admin_user(cur, _admin_user_values(payload))


@app.post("/api/admin/users/bulk", status_code=201)
def create_admin_users(
    payload: AdminUserBulkCreate,
    user: Annotated[dict, Depends(require_permission(30))],
):
    if not payload.users:
        raise HTTPException(status_code=422, detail="At least one staff account is required")
    values = [_admin_user_values(item) for item in payload.users]
    if len({item[0] for item in values}) != len(values) or len({item[2] for item in values}) != len(values):
        raise HTTPException(status_code=422, detail="The upload has duplicate staff IDs or emails")
    with get_conn() as conn:
        with conn.cursor() as cur:
            accounts = [_insert_admin_user(cur, item) for item in values]
    return {"accounts": accounts}


@app.patch("/api/admin/users/{user_id}")
def update_admin_user_status(
    user_id: int,
    payload: AdminUserUpdate,
    user: Annotated[dict, Depends(require_permission(30))],
):
    if payload.is_active is None and payload.role_name is None:
        raise HTTPException(status_code=422, detail="Provide an account status or role change")
    if user_id == user["user_id"] and payload.is_active is False:
        raise HTTPException(status_code=409, detail="You cannot deactivate your own account")
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT role_id FROM app_user WHERE user_id = %s FOR UPDATE", (user_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Staff account not found")
            updates: list[str] = []
            values: list = []
            if payload.is_active is not None:
                updates.append("is_active = %s")
                values.append(payload.is_active)
            if payload.role_name is not None:
                role_name = payload.role_name.strip().lower()
                cur.execute("SELECT role_id FROM role WHERE role_name = %s", (role_name,))
                role = cur.fetchone()
                if not role:
                    raise HTTPException(status_code=422, detail="Role must be management, coordinator, or lecturer")
                if user_id == user["user_id"] and role_name != "management":
                    raise HTTPException(status_code=409, detail="You cannot remove your own Management role")
                if role_name != "coordinator":
                    cur.execute("SELECT 1 FROM unit_offering WHERE coordinator_id = %s LIMIT 1", (user_id,))
                    if cur.fetchone():
                        raise HTTPException(status_code=409, detail="Reassign this person's coordinator offerings before changing their role")
                if role_name not in {"coordinator", "lecturer"}:
                    cur.execute("SELECT 1 FROM offering_lecturer WHERE lecturer_id = %s LIMIT 1", (user_id,))
                    if cur.fetchone():
                        raise HTTPException(status_code=409, detail="Remove this person's lecturer assignments before changing their role")
                updates.append("role_id = %s")
                values.append(role["role_id"])
            values.append(user_id)
            cur.execute(f"UPDATE app_user SET {', '.join(updates)} WHERE user_id = %s", values)
    return {"status": "updated"}


_SEMESTER_STATUSES = {"planning", "active", "archived"}
_OFFERING_STATUSES = {"draft", "active", "discontinued"}
_STUDENT_CODE_PATTERN = re.compile(r"^\d{8,9}$")
_MAX_UPLOAD_BYTES = 10 * 1024 * 1024
_MAX_UPLOAD_ROWS = 20_000
_MAX_RAW_MARK = Decimal("9999.99")


def _normalise_headers(raw_headers: list[str]) -> list[str]:
    """Keep duplicate Moodle headings usable by giving later copies a suffix."""
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, raw_header in enumerate(raw_headers, start=1):
        base = raw_header.strip() or f"Column {index}"
        seen[base] = seen.get(base, 0) + 1
        headers.append(base if seen[base] == 1 else f"{base} [{seen[base]}]")
    return headers


async def _read_csv_upload(file: UploadFile) -> tuple[str, list[str], list[tuple[int, dict[str, str]]]]:
    if not file.filename or not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=422, detail="Upload a Moodle CSV file")
    content = await file.read()
    if len(content) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="CSV files must be 10 MB or smaller")
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=422, detail="CSV must use UTF-8 encoding") from exc

    reader = csv.reader(io.StringIO(text, newline=""))
    try:
        raw_headers = next(reader)
    except StopIteration as exc:
        raise HTTPException(status_code=422, detail="CSV has no header row") from exc
    headers = _normalise_headers(raw_headers)
    if not any(header.strip() for header in raw_headers):
        raise HTTPException(status_code=422, detail="CSV has no usable header row")

    rows: list[tuple[int, dict[str, str]]] = []
    for row_number, values in enumerate(reader, start=2):
        if not any(cell.strip() for cell in values):
            continue
        if len(rows) >= _MAX_UPLOAD_ROWS:
            raise HTTPException(status_code=413, detail="CSV has too many rows")
        padded = values[: len(headers)] + [""] * max(0, len(headers) - len(values))
        rows.append((row_number, dict(zip(headers, padded, strict=True))))
    return file.filename, headers, rows


def _require_columns(headers: list[str], *columns: str) -> None:
    missing = [column for column in columns if column not in headers]
    if missing:
        raise HTTPException(status_code=422, detail=f"Selected CSV column was not found: {', '.join(missing)}")


def _admin_context_payload() -> dict:
    periods = fetch_all(
        """
        SELECT s.semester_id, s.year, s.period, s.start_date, s.end_date, s.status,
               (SELECT COUNT(*) FROM unit_offering o WHERE o.semester_id = s.semester_id) AS offering_count,
               (SELECT COUNT(*) FROM enrollment e JOIN unit_offering o ON o.offering_id = e.offering_id
                WHERE o.semester_id = s.semester_id) AS student_count,
               (SELECT COUNT(DISTINCT staff_id) FROM (
                    SELECT o.coordinator_id AS staff_id FROM unit_offering o WHERE o.semester_id = s.semester_id
                    UNION ALL
                    SELECT ol.lecturer_id FROM offering_lecturer ol
                    JOIN unit_offering o ON o.offering_id = ol.offering_id
                    WHERE o.semester_id = s.semester_id
                ) assignments) AS staff_count
        FROM semester s
        ORDER BY s.year DESC, s.period DESC
        """
    )
    offerings_rows = fetch_all(
        """
        SELECT o.offering_id, o.semester_id, o.unit_id, o.coordinator_id, o.status,
               o.handbook_url, o.last_scraped_at,
               u.unit_code, u.unit_name, replacement.unit_code AS replacement_unit_code,
               ARRAY_AGG(DISTINCT p.program_id) AS program_ids,
               ARRAY_AGG(DISTINCT p.program_code ORDER BY p.program_code) AS program_codes,
               ARRAY_AGG(DISTINCT p.program_name ORDER BY p.program_name) AS program_names,
               s.year, s.period,
               coordinator.full_name AS coordinator_name,
               ARRAY(SELECT ol.lecturer_id FROM offering_lecturer ol
                     WHERE ol.offering_id = o.offering_id ORDER BY ol.lecturer_id) AS lecturer_ids,
               (SELECT COUNT(*) FROM enrollment e WHERE e.offering_id = o.offering_id) AS student_count,
               (SELECT COUNT(*) FROM grade_upload_batch b
                WHERE b.offering_id = o.offering_id AND b.status = 'committed') AS committed_grade_upload_count
        FROM unit_offering o
        JOIN unit u ON u.unit_id = o.unit_id
        JOIN offering_program op ON op.offering_id = o.offering_id
        JOIN program p ON p.program_id = op.program_id
        JOIN semester s ON s.semester_id = o.semester_id
        JOIN app_user coordinator ON coordinator.user_id = o.coordinator_id
        LEFT JOIN unit replacement ON replacement.unit_id = o.replaced_by_unit_id
        GROUP BY o.offering_id, o.semester_id, o.unit_id, o.coordinator_id, o.status,
                 o.handbook_url, o.last_scraped_at, u.unit_code, u.unit_name,
                 replacement.unit_code, s.year, s.period, coordinator.full_name
        ORDER BY s.year DESC, s.period DESC, u.unit_code
        """
    )
    staff = fetch_all(
        """
        SELECT u.user_id, u.staff_id, u.full_name, u.email, u.is_active, u.must_change_password,
               r.role_name, r.permission_level
        FROM app_user u
        JOIN role r ON r.role_id = u.role_id
        ORDER BY u.full_name
        """
    )
    programs = fetch_all("SELECT program_id, program_code, program_name FROM program ORDER BY program_code")
    enrollment_batches = fetch_all(
        """
        SELECT b.enrollment_upload_batch_id, b.offering_id, b.original_filename, b.row_count,
               b.accepted_count, b.issue_count, b.status, b.uploaded_at,
               u.unit_code, s.year, s.period, uploader.full_name AS uploaded_by_name
        FROM enrollment_upload_batch b
        JOIN unit_offering o ON o.offering_id = b.offering_id
        JOIN unit u ON u.unit_id = o.unit_id
        JOIN semester s ON s.semester_id = o.semester_id
        JOIN app_user uploader ON uploader.user_id = b.uploaded_by
        ORDER BY b.uploaded_at DESC
        LIMIT 50
        """
    )
    return {
        "periods": periods,
        "offerings": offerings_rows,
        "staff": staff,
        "programs": programs,
        "enrollment_batches": enrollment_batches,
    }


def _validate_period(payload: AdminSemesterCreate | AdminSemesterUpdate) -> None:
    if payload.status not in _SEMESTER_STATUSES:
        raise HTTPException(status_code=422, detail="Status must be planning, active, or archived")
    if payload.start_date and payload.end_date and payload.start_date > payload.end_date:
        raise HTTPException(status_code=422, detail="Teaching end date must be after the start date")


def _validate_offering_status(status: str) -> None:
    if status not in _OFFERING_STATUSES:
        raise HTTPException(status_code=422, detail="Offering status must be draft, active, or discontinued")


def _validate_offering_staff(cur, coordinator_id: int, lecturer_ids: list[int]) -> list[int]:
    staff_ids = [coordinator_id, *lecturer_ids]
    rows = []
    if staff_ids:
        cur.execute(
            """
            SELECT u.user_id, u.is_active, r.role_name
            FROM app_user u JOIN role r ON r.role_id = u.role_id
            WHERE u.user_id = ANY(%s)
            """,
            (list(set(staff_ids)),),
        )
        rows = cur.fetchall()
    staff_by_id = {row["user_id"]: row for row in rows}
    coordinator = staff_by_id.get(coordinator_id)
    if not coordinator or not coordinator["is_active"] or coordinator["role_name"] != "coordinator":
        raise HTTPException(status_code=422, detail="Coordinator must be an active coordinator account")
    cleaned_lecturers = list(dict.fromkeys(lecturer_id for lecturer_id in lecturer_ids if lecturer_id != coordinator_id))
    for lecturer_id in cleaned_lecturers:
        lecturer = staff_by_id.get(lecturer_id)
        if not lecturer or not lecturer["is_active"] or lecturer["role_name"] not in {"coordinator", "lecturer"}:
            raise HTTPException(status_code=422, detail="Lecturers must be active coordinator or lecturer accounts")
    return cleaned_lecturers


def _save_offering_staff(cur, offering_id: int, coordinator_id: int, lecturer_ids: list[int]) -> None:
    cur.execute("UPDATE unit_offering SET coordinator_id = %s WHERE offering_id = %s", (coordinator_id, offering_id))
    cur.execute("DELETE FROM offering_lecturer WHERE offering_id = %s", (offering_id,))
    for lecturer_id in lecturer_ids:
        cur.execute(
            "INSERT INTO offering_lecturer (offering_id, lecturer_id) VALUES (%s, %s)",
            (offering_id, lecturer_id),
        )


def _validate_program_ids(cur, program_ids: list[int]) -> list[int]:
    cleaned = list(dict.fromkeys(program_ids))
    if not cleaned:
        raise HTTPException(status_code=422, detail="Select at least one program")
    cur.execute("SELECT program_id FROM program WHERE program_id = ANY(%s)", (cleaned,))
    found = {row["program_id"] for row in cur.fetchall()}
    if found != set(cleaned):
        raise HTTPException(status_code=422, detail="Program not found")
    return cleaned


def _save_offering_programs(cur, offering_id: int, program_ids: list[int]) -> None:
    cur.execute("DELETE FROM offering_program WHERE offering_id = %s", (offering_id,))
    for program_id in program_ids:
        cur.execute(
            "INSERT INTO offering_program (offering_id, program_id) VALUES (%s, %s)",
            (offering_id, program_id),
        )


@app.get("/api/admin/context")
def admin_context(user: Annotated[dict, Depends(require_permission(30))]):
    return _admin_context_payload()


@app.post("/api/admin/periods", status_code=201)
def create_admin_period(
    payload: AdminSemesterCreate,
    user: Annotated[dict, Depends(require_permission(30))],
):
    _validate_period(payload)
    period = payload.period.strip().upper()
    if period not in {"S1", "S2"}:
        raise HTTPException(status_code=422, detail="Period must be S1 or S2")
    if not 2020 <= payload.year <= 2100:
        raise HTTPException(status_code=422, detail="Year must be between 2020 and 2100")
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM semester WHERE year = %s AND period = %s", (payload.year, period))
            if cur.fetchone():
                raise HTTPException(status_code=409, detail=f"{payload.year} {period} already exists")
            if payload.status == "active":
                cur.execute("UPDATE semester SET status = 'archived' WHERE status = 'active'")
            cur.execute(
                """
                INSERT INTO semester (year, period, start_date, end_date, status)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING semester_id
                """,
                (payload.year, period, payload.start_date, payload.end_date, payload.status),
            )
            semester_id = cur.fetchone()["semester_id"]
    return {"semester_id": semester_id, "status": "created"}


@app.patch("/api/admin/periods/{semester_id}")
def update_admin_period(
    semester_id: int,
    payload: AdminSemesterUpdate,
    user: Annotated[dict, Depends(require_permission(30))],
):
    _validate_period(payload)
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM semester WHERE semester_id = %s", (semester_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Academic period not found")
            if payload.status == "active":
                cur.execute("UPDATE semester SET status = 'archived' WHERE status = 'active' AND semester_id <> %s", (semester_id,))
            cur.execute(
                """
                UPDATE semester
                SET start_date = %s, end_date = %s, status = %s
                WHERE semester_id = %s
                """,
                (payload.start_date, payload.end_date, payload.status, semester_id),
            )
    return {"status": "updated"}


@app.post("/api/admin/offerings", status_code=201)
def create_admin_offering(
    payload: AdminOfferingCreate,
    user: Annotated[dict, Depends(require_permission(30))],
):
    _validate_offering_status(payload.status)
    unit_code = payload.unit_code.strip().upper()
    unit_name = payload.unit_name.strip()
    if not re.fullmatch(r"[A-Z]{3}\d{4}", unit_code) or len(unit_name) < 3:
        raise HTTPException(status_code=422, detail="Use a valid unit code and unit name")
    with get_conn() as conn:
        with conn.cursor() as cur:
            lecturer_ids = _validate_offering_staff(cur, payload.coordinator_id, payload.lecturer_ids)
            program_ids = _validate_program_ids(cur, payload.program_ids)
            cur.execute("SELECT 1 FROM semester WHERE semester_id = %s", (payload.semester_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=422, detail="Academic period not found")
            cur.execute(
                """
                INSERT INTO unit (unit_code, unit_name)
                VALUES (%s, %s)
                ON CONFLICT (unit_code) DO UPDATE SET unit_name = EXCLUDED.unit_name
                RETURNING unit_id
                """,
                (unit_code, unit_name),
            )
            unit_id = cur.fetchone()["unit_id"]
            cur.execute(
                """
                SELECT 1 FROM unit_offering
                WHERE unit_id = %s AND semester_id = %s
                """,
                (unit_id, payload.semester_id),
            )
            if cur.fetchone():
                raise HTTPException(status_code=409, detail="This unit offering already exists for the selected period")
            cur.execute(
                """
                INSERT INTO unit_offering (unit_id, semester_id, coordinator_id, status)
                VALUES (%s, %s, %s, %s)
                RETURNING offering_id
                """,
                (unit_id, payload.semester_id, payload.coordinator_id, payload.status),
            )
            offering_id = cur.fetchone()["offering_id"]
            _save_offering_programs(cur, offering_id, program_ids)
            _save_offering_staff(cur, offering_id, payload.coordinator_id, lecturer_ids)
    return {"offering_id": offering_id, "status": "created"}


@app.patch("/api/admin/offerings/{offering_id}")
def update_admin_offering(
    offering_id: int,
    payload: AdminOfferingUpdate,
    user: Annotated[dict, Depends(require_permission(30))],
):
    _validate_offering_status(payload.status)
    unit_code = payload.unit_code.strip().upper()
    unit_name = payload.unit_name.strip()
    if not re.fullmatch(r"[A-Z]{3}\d{4}", unit_code) or len(unit_name) < 3:
        raise HTTPException(status_code=422, detail="Use a valid unit code and unit name")
    replacement_code = (payload.replacement_unit_code or "").strip().upper()
    replacement_name = (payload.replacement_unit_name or "").strip()
    if replacement_code and not re.fullmatch(r"[A-Z]{3}\d{4}", replacement_code):
        raise HTTPException(status_code=422, detail="Replacement unit code must look like FIT3161")
    with get_conn() as conn:
        with conn.cursor() as cur:
            lecturer_ids = _validate_offering_staff(cur, payload.coordinator_id, payload.lecturer_ids)
            program_ids = _validate_program_ids(cur, payload.program_ids)
            cur.execute("SELECT unit_id FROM unit_offering WHERE offering_id = %s FOR UPDATE", (offering_id,))
            offering = cur.fetchone()
            if not offering:
                raise HTTPException(status_code=404, detail="Unit offering not found")
            cur.execute("SELECT unit_id FROM unit WHERE unit_code = %s AND unit_id <> %s", (unit_code, offering["unit_id"]))
            if cur.fetchone():
                raise HTTPException(status_code=409, detail="Another unit already uses this unit code")
            replacement_unit_id = None
            if replacement_code:
                cur.execute("SELECT unit_id FROM unit WHERE unit_code = %s", (replacement_code,))
                replacement = cur.fetchone()
                if replacement:
                    replacement_unit_id = replacement["unit_id"]
                elif replacement_name:
                    cur.execute(
                        "INSERT INTO unit (unit_code, unit_name) VALUES (%s, %s) RETURNING unit_id",
                        (replacement_code, replacement_name),
                    )
                    replacement_unit_id = cur.fetchone()["unit_id"]
                else:
                    raise HTTPException(status_code=422, detail="Enter the replacement unit name when adding a new replacement code")
            cur.execute(
                "UPDATE unit SET unit_code = %s, unit_name = %s WHERE unit_id = %s",
                (unit_code, unit_name, offering["unit_id"]),
            )
            cur.execute(
                "UPDATE unit_offering SET status = %s, replaced_by_unit_id = %s WHERE offering_id = %s",
                (payload.status, replacement_unit_id, offering_id),
            )
            _save_offering_programs(cur, offering_id, program_ids)
            _save_offering_staff(cur, offering_id, payload.coordinator_id, lecturer_ids)
    return {"status": "updated"}


def _validate_enrolment_rows(
    rows: list[tuple[int, dict[str, str]]],
    student_code_column: str,
    full_name_column: str,
) -> tuple[list[dict], int]:
    issues: list[dict] = []
    accepted_count = 0
    seen_codes: set[str] = set()
    for row_number, row in rows:
        student_code = row[student_code_column].strip().replace(" ", "")
        full_name = row[full_name_column].strip()
        if not student_code:
            issues.append({"row": row_number, "severity": "error", "message": "Missing student ID"})
        elif not _STUDENT_CODE_PATTERN.fullmatch(student_code):
            issues.append({"row": row_number, "severity": "error", "message": "Student ID must have 8 or 9 digits"})
        elif student_code in seen_codes:
            issues.append({"row": row_number, "severity": "error", "message": "Duplicate student ID in this file"})
        elif len(full_name) < 3:
            issues.append({"row": row_number, "severity": "error", "message": "Missing student name"})
        elif len(full_name) > 150:
            issues.append({"row": row_number, "severity": "error", "message": "Student name is too long"})
        else:
            accepted_count += 1
        seen_codes.add(student_code)
    return issues, accepted_count


@app.post("/api/admin/enrolments/inspect")
async def inspect_enrolment_upload(
    user: Annotated[dict, Depends(require_permission(30))],
    file: UploadFile = File(...),
):
    filename, headers, rows = await _read_csv_upload(file)
    return {"filename": filename, "headers": headers, "row_count": len(rows)}


@app.post("/api/admin/enrolments/preview")
async def preview_enrolment_upload(
    user: Annotated[dict, Depends(require_permission(30))],
    offering_id: int = Form(...),
    student_code_column: str = Form(...),
    full_name_column: str = Form(...),
    file: UploadFile = File(...),
):
    if not fetch_one("SELECT 1 FROM unit_offering WHERE offering_id = %s", (offering_id,)):
        raise HTTPException(status_code=404, detail="Unit offering not found")
    filename, headers, rows = await _read_csv_upload(file)
    _require_columns(headers, student_code_column, full_name_column)
    issues, accepted_count = _validate_enrolment_rows(rows, student_code_column, full_name_column)
    return {
        "filename": filename,
        "row_count": len(rows),
        "accepted_count": accepted_count,
        "issues": issues,
        "status": "valid" if not issues else "needs_review",
    }


@app.post("/api/admin/enrolments/commit")
async def commit_enrolment_upload(
    user: Annotated[dict, Depends(require_permission(30))],
    offering_id: int = Form(...),
    student_code_column: str = Form(...),
    full_name_column: str = Form(...),
    file: UploadFile = File(...),
):
    filename, headers, rows = await _read_csv_upload(file)
    _require_columns(headers, student_code_column, full_name_column)
    issues, accepted_count = _validate_enrolment_rows(rows, student_code_column, full_name_column)
    if any(issue["severity"] == "error" for issue in issues):
        raise HTTPException(status_code=422, detail="Fix all student-list errors before committing")
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT 1 FROM unit_offering WHERE offering_id = %s FOR UPDATE",
                (offering_id,),
            )
            offering = cur.fetchone()
            if not offering:
                raise HTTPException(status_code=404, detail="Unit offering not found")
            cur.execute("SELECT program_id FROM offering_program WHERE offering_id = %s", (offering_id,))
            offering_program_ids = [row["program_id"] for row in cur.fetchall()]
            # A student's program is only unambiguous when the offering serves a single program.
            student_program_id = offering_program_ids[0] if len(offering_program_ids) == 1 else None
            for _, row in rows:
                student_code = row[student_code_column].strip().replace(" ", "")
                full_name = row[full_name_column].strip()
                cur.execute(
                    """
                    INSERT INTO student (student_code, full_name, program_id)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (student_code) DO UPDATE SET full_name = EXCLUDED.full_name
                    RETURNING student_id
                    """,
                    (student_code, full_name, student_program_id),
                )
                student_id = cur.fetchone()["student_id"]
                cur.execute(
                    "INSERT INTO enrollment (student_id, offering_id) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                    (student_id, offering_id),
                )
            cur.execute(
                """
                INSERT INTO enrollment_upload_batch (
                    offering_id, uploaded_by, original_filename, row_count, accepted_count, issue_count, status
                ) VALUES (%s, %s, %s, %s, %s, %s, 'committed')
                RETURNING enrollment_upload_batch_id
                """,
                (offering_id, user["user_id"], filename, len(rows), accepted_count, len(issues)),
            )
            batch_id = cur.fetchone()["enrollment_upload_batch_id"]
    return {"status": "committed", "batch_id": batch_id, "accepted_count": accepted_count}


def _grade_column_mappings(raw_mapping: str, headers: list[str], assessment_by_id: dict[int, dict]) -> list[dict]:
    try:
        parsed = json.loads(raw_mapping)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=422, detail="Grade column mapping is invalid") from exc
    if not isinstance(parsed, list) or not parsed:
        raise HTTPException(status_code=422, detail="Map at least one assessment column")
    mappings: list[dict] = []
    used_columns: set[str] = set()
    used_assessments: set[int] = set()
    for item in parsed:
        if not isinstance(item, dict):
            raise HTTPException(status_code=422, detail="Grade column mapping is invalid")
        try:
            assessment_id = int(item["assessment_id"])
            csv_column = str(item["csv_column"])
            max_mark = Decimal(str(item["max_mark"]))
        except (KeyError, ValueError, ArithmeticError) as exc:
            raise HTTPException(status_code=422, detail="Each grade column needs an assessment and maximum mark") from exc
        if assessment_id not in assessment_by_id:
            raise HTTPException(status_code=422, detail="Selected assessment does not belong to this offering")
        if csv_column not in headers:
            raise HTTPException(status_code=422, detail=f"Selected CSV column was not found: {csv_column}")
        if max_mark <= 0 or max_mark > _MAX_RAW_MARK:
            raise HTTPException(status_code=422, detail="Maximum mark must be between 0 and 9999.99")
        if csv_column in used_columns or assessment_id in used_assessments:
            raise HTTPException(status_code=422, detail="Each assessment and CSV column can only be mapped once")
        used_columns.add(csv_column)
        used_assessments.add(assessment_id)
        mappings.append({"assessment_id": assessment_id, "csv_column": csv_column, "max_mark": max_mark})
    return mappings


def _load_offering_assessments(cur, offering_id: int) -> dict[int, dict]:
    cur.execute(
        """
        SELECT assessment_id, assessment_name, weight
        FROM assessment WHERE offering_id = %s ORDER BY assessment_order
        """,
        (offering_id,),
    )
    return {row["assessment_id"]: row for row in cur.fetchall()}


@app.post("/api/grade-uploads/inspect")
async def inspect_grade_upload(
    user: Annotated[dict, Depends(require_permission(10))],
    offering_id: int = Form(...),
    file: UploadFile = File(...),
):
    ensure_offering_access(user, offering_id, min_permission_level=10)
    filename, headers, rows = await _read_csv_upload(file)
    return {"filename": filename, "headers": headers, "row_count": len(rows), "offering_id": offering_id}


@app.post("/api/grade-uploads/preview")
async def preview_grade_upload(
    user: Annotated[dict, Depends(require_permission(10))],
    offering_id: int = Form(...),
    student_code_column: str = Form(...),
    assessment_columns: str = Form(...),
    file: UploadFile = File(...),
):
    ensure_offering_access(user, offering_id, min_permission_level=10)
    filename, headers, rows = await _read_csv_upload(file)
    _require_columns(headers, student_code_column)
    with get_conn() as conn:
        with conn.cursor() as cur:
            assessment_by_id = _load_offering_assessments(cur, offering_id)
            if not assessment_by_id:
                raise HTTPException(status_code=422, detail="Confirm the assessment setup before importing grades")
            mappings = _grade_column_mappings(assessment_columns, headers, assessment_by_id)
            cur.execute(
                """
                SELECT e.enrollment_id, s.student_id, s.student_code, s.full_name
                FROM enrollment e JOIN student s ON s.student_id = e.student_id
                WHERE e.offering_id = %s
                """,
                (offering_id,),
            )
            enrolled_by_code = {row["student_code"]: row for row in cur.fetchall()}
            cur.execute(
                """
                INSERT INTO grade_upload_batch (offering_id, uploaded_by, original_filename, status)
                VALUES (%s, %s, %s, 'draft')
                RETURNING upload_batch_id
                """,
                (offering_id, user["user_id"], filename),
            )
            batch_id = cur.fetchone()["upload_batch_id"]
            cur.execute(
                """
                INSERT INTO grade_upload_column_mapping (upload_batch_id, csv_column_name, system_field)
                VALUES (%s, %s, 'student_code')
                """,
                (batch_id, student_code_column),
            )
            for mapping in mappings:
                cur.execute(
                    """
                    INSERT INTO grade_upload_column_mapping (
                        upload_batch_id, csv_column_name, system_field, assessment_id, max_mark
                    ) VALUES (%s, %s, 'raw_mark', %s, %s)
                    """,
                    (batch_id, mapping["csv_column"], mapping["assessment_id"], mapping["max_mark"]),
                )

            issues: list[dict] = []
            seen_codes: set[str] = set()
            matched_count = 0
            for row_number, row in rows:
                student_code = row[student_code_column].strip().replace(" ", "")
                row_issues: list[tuple[str, str, str]] = []
                cells: list[tuple[int, Decimal, Decimal]] = []
                if not student_code:
                    row_issues.append(("missing_student_id", "error", "Missing student ID"))
                elif student_code in seen_codes:
                    row_issues.append(("duplicate_student_id", "error", "Duplicate student ID in this file"))
                elif student_code not in enrolled_by_code:
                    row_issues.append(("unmatched_student", "error", "Student ID is not enrolled in this offering"))
                else:
                    for mapping in mappings:
                        try:
                            mark = parse_mark(row[mapping["csv_column"]])
                        except ValueError:
                            row_issues.append(
                                ("invalid_mark", "error", f"{mapping['csv_column']} is not a numeric mark")
                            )
                            continue
                        if mark is None:
                            continue
                        if mark > _MAX_RAW_MARK:
                            row_issues.append(
                                ("mark_out_of_range", "error", f"{mapping['csv_column']} is larger than the supported mark range")
                            )
                            continue
                        try:
                            weighted_score(mark, mapping["max_mark"], assessment_by_id[mapping["assessment_id"]]["weight"])
                        except ValueError:
                            row_issues.append(
                                ("mark_out_of_range", "error", f"{mapping['csv_column']} must be between 0 and {mapping['max_mark']}")
                            )
                            continue
                        cells.append((mapping["assessment_id"], mark, mapping["max_mark"]))
                    if not cells and not row_issues:
                        row_issues.append(("no_marks", "warning", "No grade values were found for the mapped assessment columns"))
                if student_code:
                    seen_codes.add(student_code)
                severity = "error" if any(issue[1] == "error" for issue in row_issues) else "warning" if row_issues else "valid"
                matched_student_id = enrolled_by_code.get(student_code, {}).get("student_id")
                cur.execute(
                    """
                    INSERT INTO grade_upload_row (upload_batch_id, row_number, student_code_raw, matched_student_id, status)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING upload_row_id
                    """,
                    (batch_id, row_number, student_code or None, matched_student_id, severity),
                )
                upload_row_id = cur.fetchone()["upload_row_id"]
                for issue_type, issue_severity, message in row_issues:
                    cur.execute(
                        """
                        INSERT INTO grade_upload_issue (upload_batch_id, upload_row_id, issue_type, severity, message)
                        VALUES (%s, %s, %s, %s, %s)
                        """,
                        (batch_id, upload_row_id, issue_type, issue_severity, message),
                    )
                    issues.append({"row": row_number, "severity": issue_severity, "message": message})
                if severity != "error":
                    for assessment_id, raw_mark, max_mark in cells:
                        cur.execute(
                            """
                            INSERT INTO grade_upload_cell (upload_row_id, assessment_id, raw_mark, max_mark)
                            VALUES (%s, %s, %s, %s)
                            """,
                            (upload_row_id, assessment_id, raw_mark, max_mark),
                        )
                    if student_code in enrolled_by_code:
                        matched_count += 1

            for student_code, enrolled in enrolled_by_code.items():
                if student_code not in seen_codes:
                    message = f"Enrolled student {student_code} is absent from this CSV"
                    cur.execute(
                        """
                        INSERT INTO grade_upload_issue (upload_batch_id, issue_type, severity, message)
                        VALUES (%s, 'missing_enrolled_student', 'warning', %s)
                        """,
                        (batch_id, message),
                    )
                    issues.append({"row": None, "severity": "warning", "message": message})
            has_errors = any(issue["severity"] == "error" for issue in issues)
            cur.execute(
                "UPDATE grade_upload_batch SET status = %s WHERE upload_batch_id = %s",
                ("rejected" if has_errors else "validated", batch_id),
            )
    return {
        "upload_batch_id": batch_id,
        "filename": filename,
        "row_count": len(rows),
        "matched_count": matched_count,
        "issues": issues,
        "status": "needs_review" if has_errors else "valid",
    }


def _recalculate_attainment(cur, offering_id: int) -> int:
    cur.execute("DELETE FROM student_ulo_attainment WHERE offering_id = %s", (offering_id,))
    cur.execute("DELETE FROM cohort_ulo_attainment WHERE offering_id = %s", (offering_id,))
    cur.execute(
        """
        WITH calculated AS (
            SELECT
                e.enrollment_id,
                au.offering_ulo_id,
                SUM(au.allocated_weight) AS total_available_weight,
                SUM(COALESCE((sg.raw_mark / NULLIF(sg.max_mark, 0)) * au.allocated_weight, 0)) AS achieved_weight
            FROM enrollment e
            JOIN assessment_ulo au ON au.offering_id = e.offering_id
            LEFT JOIN student_grade sg ON sg.enrollment_id = e.enrollment_id
                AND sg.assessment_id = au.assessment_id
            WHERE e.offering_id = %s AND au.allocated_weight > 0
            GROUP BY e.enrollment_id, au.offering_ulo_id
        )
        INSERT INTO student_ulo_attainment (
            offering_id, enrollment_id, offering_ulo_id, total_available_weight,
            achieved_weight, attainment_pct, is_achieved, calculated_at
        )
        SELECT %s, enrollment_id, offering_ulo_id, total_available_weight, achieved_weight,
               ROUND((achieved_weight / NULLIF(total_available_weight, 0)) * 100, 2),
               (achieved_weight / NULLIF(total_available_weight, 0)) * 100 >= 50,
               CURRENT_TIMESTAMP
        FROM calculated
        WHERE total_available_weight > 0
        """,
        (offering_id, offering_id),
    )
    attainment_count = cur.rowcount
    cur.execute(
        """
        SELECT COUNT(*) AS enrolled_count FROM enrollment WHERE offering_id = %s
        """,
        (offering_id,),
    )
    enrolled_count = cur.fetchone()["enrolled_count"]
    cur.execute(
        """
        INSERT INTO cohort_ulo_attainment (
            offering_id, offering_ulo_id, enrolled_count, achieved_count,
            average_attainment_pct, pass_rate_pct, calculated_at
        )
        SELECT
            %s,
            ou.offering_ulo_id,
            %s,
            COUNT(sua.attainment_id) FILTER (WHERE sua.is_achieved),
            COALESCE(ROUND(AVG(sua.attainment_pct), 2), 0),
            CASE WHEN %s = 0 THEN 0
                 ELSE ROUND((COUNT(sua.attainment_id) FILTER (WHERE sua.is_achieved))::numeric / %s * 100, 2)
            END,
            CURRENT_TIMESTAMP
        FROM offering_ulo ou
        LEFT JOIN student_ulo_attainment sua ON sua.offering_ulo_id = ou.offering_ulo_id
            AND sua.offering_id = %s
        WHERE ou.offering_id = %s
        GROUP BY ou.offering_ulo_id
        """,
        (offering_id, enrolled_count, enrolled_count, enrolled_count, offering_id, offering_id),
    )
    return attainment_count


@app.post("/api/grade-uploads/{upload_batch_id}/commit")
def commit_grade_upload(
    upload_batch_id: int,
    user: Annotated[dict, Depends(require_permission(10))],
):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT upload_batch_id, offering_id, status
                FROM grade_upload_batch WHERE upload_batch_id = %s FOR UPDATE
                """,
                (upload_batch_id,),
            )
            batch = cur.fetchone()
            if not batch:
                raise HTTPException(status_code=404, detail="Grade upload preview not found")
            ensure_offering_access(user, batch["offering_id"], min_permission_level=10)
            if batch["status"] != "validated":
                raise HTTPException(status_code=409, detail="This grade upload has errors and cannot be committed")
            cur.execute(
                """
                SELECT COUNT(*) AS count FROM grade_upload_issue
                WHERE upload_batch_id = %s AND severity = 'error'
                """,
                (upload_batch_id,),
            )
            if cur.fetchone()["count"]:
                raise HTTPException(status_code=409, detail="Resolve upload errors before committing")
            cur.execute(
                """
                SELECT e.enrollment_id, c.assessment_id, c.raw_mark, c.max_mark, a.weight, r.upload_row_id
                FROM grade_upload_cell c
                JOIN grade_upload_row r ON r.upload_row_id = c.upload_row_id
                JOIN enrollment e ON e.student_id = r.matched_student_id AND e.offering_id = %s
                JOIN assessment a ON a.assessment_id = c.assessment_id AND a.offering_id = %s
                WHERE r.upload_batch_id = %s AND r.status IN ('valid', 'warning') AND c.status = 'valid'
                """,
                (batch["offering_id"], batch["offering_id"], upload_batch_id),
            )
            grade_rows = cur.fetchall()
            for grade in grade_rows:
                score = weighted_score(grade["raw_mark"], grade["max_mark"], grade["weight"])
                cur.execute(
                    """
                    INSERT INTO student_grade (
                        offering_id, enrollment_id, assessment_id, upload_batch_id, source_row_id,
                        raw_mark, max_mark, weighted_score
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (enrollment_id, assessment_id) DO UPDATE SET
                        upload_batch_id = EXCLUDED.upload_batch_id,
                        source_row_id = EXCLUDED.source_row_id,
                        raw_mark = EXCLUDED.raw_mark,
                        max_mark = EXCLUDED.max_mark,
                        weighted_score = EXCLUDED.weighted_score,
                        uploaded_at = CURRENT_TIMESTAMP
                    """,
                    (
                        batch["offering_id"], grade["enrollment_id"], grade["assessment_id"], upload_batch_id,
                        grade["upload_row_id"], grade["raw_mark"], grade["max_mark"], score,
                    ),
                )
            cur.execute(
                """
                UPDATE grade_upload_row SET status = 'committed'
                WHERE upload_batch_id = %s AND status IN ('valid', 'warning')
                """,
                (upload_batch_id,),
            )
            attainment_count = _recalculate_attainment(cur, batch["offering_id"])
            cur.execute(
                """
                UPDATE grade_upload_batch
                SET status = 'committed', committed_at = CURRENT_TIMESTAMP
                WHERE upload_batch_id = %s
                """,
                (upload_batch_id,),
            )
    return {
        "status": "committed",
        "grades_saved": len(grade_rows),
        "attainment_records": attainment_count,
    }


# ---------------------------------------------------------------------------
# Teaching staff roster (Lecture / Tutorial / Laboratory) — e.g. the SoIT
# "Lecture and Tutor List" spreadsheet. This is a reporting/compliance roster,
# separate from offering_lecturer (which controls dashboard login access).
# ---------------------------------------------------------------------------

_ROSTER_ROLE_COLUMNS = [
    ("lecture", 4, 5),
    ("tutorial", 6, 7),
    ("laboratory", 8, 9),
]
_NOT_APPLICABLE = {"", "not applicable", "n/a", "na"}


def _roster_cell(row: tuple, index: int) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _parse_staffing_roster(content: bytes) -> list[dict]:
    """Parse a block-structured roster: one unit spans 1+ rows, later rows in
    the same block leave 'Unit Code' blank and only add more staff."""
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="Could not read the roster spreadsheet") from exc
    sheet = workbook.worksheets[0]
    units: list[dict] = []
    current: dict | None = None
    for row in sheet.iter_rows(min_row=4, values_only=True):
        unit_code = _roster_cell(row, 2).upper()
        if unit_code:
            programme_raw = _roster_cell(row, 1)
            current = {
                "unit_code": unit_code,
                "unit_name": _roster_cell(row, 3),
                "programme_codes": [item.strip().upper() for item in programme_raw.split("/") if item.strip()],
                "staffing": [],
            }
            units.append(current)
        if current is None:
            continue
        for role_type, name_col, email_col in _ROSTER_ROLE_COLUMNS:
            name = _roster_cell(row, name_col)
            if name.lower() in _NOT_APPLICABLE:
                continue
            email = _roster_cell(row, email_col).lower() or None
            current["staffing"].append({"role_type": role_type, "name": name, "email": email})
    return units


@app.post("/api/admin/staffing/roster-import")
async def import_staffing_roster(
    user: Annotated[dict, Depends(require_permission(30))],
    semester_id: int = Form(...),
    file: UploadFile = File(...),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="Upload an .xlsx roster file")
    content = await file.read()
    if len(content) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="Roster files must be 10 MB or smaller")
    units = _parse_staffing_roster(content)
    if not units:
        raise HTTPException(status_code=422, detail="No units were found in this roster file")

    warnings: list[str] = []
    unmatched_units: list[dict] = []
    matched_offerings = 0
    staffing_rows_created = 0

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT 1 FROM semester WHERE semester_id = %s", (semester_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=422, detail="Academic period not found")
            for unit in units:
                cur.execute(
                    """
                    SELECT o.offering_id FROM unit_offering o
                    JOIN unit u ON u.unit_id = o.unit_id
                    WHERE u.unit_code = %s AND o.semester_id = %s
                    """,
                    (unit["unit_code"], semester_id),
                )
                offering = cur.fetchone()
                if not offering:
                    unmatched_units.append(
                        {"unit_code": unit["unit_code"], "unit_name": unit["unit_name"], "programme_codes": unit["programme_codes"]}
                    )
                    continue
                offering_id = offering["offering_id"]
                matched_offerings += 1

                if unit["programme_codes"]:
                    cur.execute(
                        "SELECT program_id, program_code FROM program WHERE UPPER(program_code) = ANY(%s)",
                        (unit["programme_codes"],),
                    )
                    found_programs = cur.fetchall()
                    found_codes = {row["program_code"].upper() for row in found_programs}
                    for code in unit["programme_codes"]:
                        if code not in found_codes:
                            warnings.append(f"{unit['unit_code']}: no matching program for '{code}'")
                    for program in found_programs:
                        cur.execute(
                            "INSERT INTO offering_program (offering_id, program_id) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                            (offering_id, program["program_id"]),
                        )

                cur.execute(
                    "DELETE FROM offering_staffing WHERE offering_id = %s AND source = 'roster_import'",
                    (offering_id,),
                )
                for entry in unit["staffing"]:
                    staff_user_id = None
                    if entry["email"]:
                        cur.execute("SELECT user_id FROM app_user WHERE LOWER(email) = %s", (entry["email"],))
                        match = cur.fetchone()
                        staff_user_id = match["user_id"] if match else None
                    cur.execute(
                        """
                        INSERT INTO offering_staffing (offering_id, role_type, staff_user_id, external_name, external_email)
                        VALUES (%s, %s, %s, %s, %s)
                        """,
                        (offering_id, entry["role_type"], staff_user_id, entry["name"], entry["email"]),
                    )
                    staffing_rows_created += 1

    return {
        "status": "imported",
        "units_in_file": len(units),
        "matched_offerings": matched_offerings,
        "staffing_rows_created": staffing_rows_created,
        "unmatched_units": unmatched_units,
        "warnings": warnings,
    }


@app.get("/api/offerings/{offering_id}/staffing")
def offering_staffing(
    offering_id: int,
    user: Annotated[dict, Depends(require_offering_access())],
):
    rows = fetch_all(
        """
        SELECT s.staffing_id, s.role_type, s.staff_user_id, s.external_name, s.external_email,
               u.full_name AS staff_full_name
        FROM offering_staffing s
        LEFT JOIN app_user u ON u.user_id = s.staff_user_id
        WHERE s.offering_id = %s
        ORDER BY s.role_type, COALESCE(u.full_name, s.external_name)
        """,
        (offering_id,),
    )
    return {"staffing": rows}
